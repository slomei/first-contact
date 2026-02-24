"""
Binary document text extraction — PDF, DOCX, XLSX.

Provides extract_text(filepath) which routes to the correct parser
based on file extension. All parsers return plain text strings.

Dependencies are optional: pdfplumber (already in requirements.txt),
python-docx, openpyxl. Missing packages produce clear error messages.
"""

import os

# Optional imports — guarded at usage sites
try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import docx as _docx
except ImportError:
    _docx = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


# Extensions this module can handle
BINARY_EXTENSIONS = {".pdf", ".docx", ".xlsx"}


def is_binary_document(filepath):
    """True if the file extension is a supported binary document format."""
    _, ext = os.path.splitext(filepath)
    return ext.lower() in BINARY_EXTENSIONS


def extract_text(filepath):
    """Extract plain text from a binary document.

    Returns the extracted text string.
    Raises ValueError if extraction fails or returns empty.
    Raises ImportError if the required library is not installed.
    """
    _, ext = os.path.splitext(filepath)
    ext = ext.lower()
    filename = os.path.basename(filepath)

    if ext == ".pdf":
        return _extract_pdf(filepath, filename)
    elif ext == ".docx":
        return _extract_docx(filepath, filename)
    elif ext == ".xlsx":
        return _extract_xlsx(filepath, filename)
    else:
        raise ValueError(f"Unsupported binary format: {ext}")


def _extract_pdf(filepath, filename):
    """Extract text from a PDF using pdfplumber."""
    if pdfplumber is None:
        raise ImportError(
            "pdfplumber is required to read PDF files. "
            "Install it with: pip install pdfplumber"
        )
    pages = []
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if text.strip():
                pages.append(text)
    result = "\n\n".join(pages)
    if not result.strip():
        raise ValueError(f"Could not extract text from {filename}")
    return result


def _extract_docx(filepath, filename):
    """Extract text from a DOCX using python-docx."""
    if _docx is None:
        raise ImportError(
            "python-docx is required to read DOCX files. "
            "Install it with: pip install python-docx"
        )
    doc = _docx.Document(filepath)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    result = "\n\n".join(paragraphs)
    if not result.strip():
        raise ValueError(f"Could not extract text from {filename}")
    return result


def _extract_xlsx(filepath, filename):
    """Extract text from an XLSX using openpyxl."""
    if openpyxl is None:
        raise ImportError(
            "openpyxl is required to read XLSX files. "
            "Install it with: pip install openpyxl"
        )
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            header = f"[Sheet: {sheet_name}]" if len(wb.sheetnames) > 1 else ""
            if header:
                rows.insert(0, header)
            sheets.append("\n".join(rows))
    wb.close()
    result = "\n\n".join(sheets)
    if not result.strip():
        raise ValueError(f"Could not extract text from {filename}")
    return result
